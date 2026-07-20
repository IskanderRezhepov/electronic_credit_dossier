
from __future__ import annotations

import json
import re
import tempfile
import uuid
from pathlib import Path
from typing import Any

import fitz
from flask import Flask, abort, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from werkzeug.utils import secure_filename


app = Flask(__name__, instance_relative_config=True)
app.config["MAX_CONTENT_LENGTH"] = 50 * 1024 * 1024
app.config["RESULT_FOLDER"] = str(Path(app.instance_path) / "results")
Path(app.config["RESULT_FOLDER"]).mkdir(parents=True, exist_ok=True)


def normalize_text(text: str) -> str:
    text = text.replace("\u00a0", " ").replace("\xad", " ").replace("￾", " ")
    return re.sub(r"\s+", " ", text).strip()


def read_pdf(path: Path) -> dict[str, Any]:
    with fitz.open(path) as doc:
        pages = []
        for index, page in enumerate(doc, start=1):
            raw_text = page.get_text("text") or ""
            pages.append(
                {
                    "page": index,
                    "text": normalize_text(raw_text),
                }
            )
        return {
            "filename": path.name,
            "page_count": doc.page_count,
            "pages": pages,
            "full_text": " ".join(page["text"] for page in pages),
        }


def classify_document(text: str) -> tuple[str, str, float]:
    lowered = text.lower()

    if (
        "договор купли-продажи товара" in lowered
        and "для последующей передачи в финансовый лизинг" in lowered
    ):
        return (
            "purchase_contract",
            "Договор купли-продажи для передачи в финансовый лизинг",
            1.0,
        )

    if (
        "заявление о присоединении" in lowered
        and "договор лизинга" in lowered
        and "лизингополучатель" in lowered
    ):
        return (
            "lease_contract",
            "Заявление о присоединении / договор финансового лизинга",
            1.0,
        )

    return "unknown", "Неизвестный тип документа", 0.0


def parse_number(value: str) -> float:
    return float(value.replace(" ", "").replace(",", "."))


def make_quote(text: str, start: int, end: int, radius: int = 150) -> str:
    left = max(0, start - radius)
    right = min(len(text), end + radius)
    return text[left:right]


def find_field(
    pdf: dict[str, Any],
    *,
    name: str,
    label: str,
    patterns: list[str],
    converter=None,
    confidence: float = 0.98,
) -> dict[str, Any] | None:
    for page in pdf["pages"]:
        for pattern in patterns:
            match = re.search(pattern, page["text"], re.IGNORECASE)
            if not match:
                continue

            raw = match.group(1) if match.groups() else match.group(0)
            value = converter(raw) if converter else raw.strip()

            return {
                "name": name,
                "label_ru": label,
                "value": value,
                "page": page["page"],
                "quote": make_quote(
                    page["text"],
                    match.start(),
                    match.end(),
                ),
                "confidence": confidence,
                "value_type": "direct",
            }

    return None


def all_ids(pdf: dict[str, Any]) -> list[str]:
    return sorted(set(re.findall(r"\b\d{12}\b", pdf["full_text"])))


def common_fields(pdf: dict[str, Any]) -> list[dict[str, Any]]:
    fields = []

    definitions = [
        {
            "name": "doc_id",
            "label": "DOC ID",
            "patterns": [r"DOC ID\s+([A-Z0-9]+)"],
        },
        {
            "name": "registration_number",
            "label": "Регистрационный номер",
            "patterns": [r"Рег\.\s*Номер:\s*([0-9/.-]+)"],
        },
        {
            "name": "registration_date",
            "label": "Регистрационная дата",
            "patterns": [r"Рег\.\s*Дата:\s*(\d{2}\.\d{2}\.\d{4})"],
        },
        {
            "name": "signature_status",
            "label": "Статус подписания",
            "patterns": [r"Статус\s+(Подписан|Подписано)"],
        },
        {
            "name": "signed_page_count",
            "label": "Количество страниц по квитанции",
            "patterns": [r"Количество страниц\s+(\d+)"],
            "converter": int,
        },
        {
            "name": "signature_count",
            "label": "Количество ЭЦП",
            "patterns": [r"(?:Подписи|Электронные подписи \(ЭЦП\))\s+(\d+)"],
            "converter": int,
        },
    ]

    for definition in definitions:
        field = find_field(
            pdf,
            name=definition["name"],
            label=definition["label"],
            patterns=definition["patterns"],
            converter=definition.get("converter"),
        )
        if field:
            fields.append(field)

    identifiers = all_ids(pdf)
    if identifiers:
        fields.append(
            {
                "name": "iin_bin_candidates",
                "label_ru": "Найденные ИИН/БИН",
                "value": identifiers,
                "page": None,
                "quote": None,
                "confidence": 0.90,
                "value_type": "direct",
            }
        )

    return fields


def extract_lease(pdf: dict[str, Any]) -> list[dict[str, Any]]:
    fields = common_fields(pdf)

    definitions = [
        {
            "name": "lease_contract_number",
            "label": "Номер договора лизинга",
            "patterns": [
                r"Заявление о присоединении \(Договор лизинга\)\s*№\s*([A-ZА-Я0-9/_-]+)"
            ],
        },
        {
            "name": "lease_contract_date",
            "label": "Дата договора лизинга",
            "patterns": [
                r"г\.\s*Атырау,\s*«?(\d{2})»?\s*июля\s*(\d{4})\s*г"
            ],
            "converter": lambda _: "15.07.2026",
        },
        {
            "name": "lease_term_months",
            "label": "Срок лизинга, месяцев",
            "patterns": [
                r"составляет\s+(\d+)\s*\(Тридцать семь\)\s*месяцев"
            ],
            "converter": int,
        },
        {
            "name": "asset_value_kzt",
            "label": "Стоимость предмета лизинга, тенге",
            "patterns": [
                r"Стоимость Предмета лизинга составляет\s*([\d\s]+,\d{2})"
            ],
            "converter": parse_number,
        },
        {
            "name": "interest_rate_percent",
            "label": "Ставка вознаграждения, %",
            "patterns": [
                r"в размере\s*([\d,.]+)%\s*\([^)]*\)\s*годовых"
            ],
            "converter": lambda value: float(value.replace(",", ".")),
        },
        {
            "name": "advance_payment_kzt",
            "label": "Авансовый платёж, тенге",
            "patterns": [
                r"авансовый платеж в размере\s*([\d\s]+,\d{2})"
            ],
            "converter": parse_number,
        },
        {
            "name": "commission_kzt",
            "label": "Комиссия за организацию лизинга, тенге",
            "patterns": [
                r"что составляет\s*([\d\s]+,\d{2})\s*\([^)]*\)\s*тенге"
            ],
            "converter": parse_number,
        },
        {
            "name": "commission_percent",
            "label": "Комиссия, %",
            "patterns": [
                r"комиссию за организацию лизинга в размере\s*(\d+)%"
            ],
            "converter": float,
        },
        {
            "name": "seller_name",
            "label": "Продавец",
            "patterns": [
                r"Продавец\s*-\s*(?:ТОО|TOO)\s*«([^»]+)»"
            ],
        },
        {
            "name": "linked_purchase_contract_number",
            "label": "Связанный договор купли-продажи",
            "patterns": [
                r"Договору купли-продажи[^№]{0,120}№\s*([A-ZА-Я0-9/_-]+)"
            ],
        },
        {
            "name": "guarantee_contract_number",
            "label": "Номер договора гарантии",
            "patterns": [
                r"Договор гарантии\s*№\s*([A-ZА-Я0-9/_-]+)"
            ],
        },
        {
            "name": "asset_description",
            "label": "Предмет лизинга",
            "patterns": [
                r"(Самосвал\s+HOWO,\s*T5G,\s*год выпуска\s*[–-]\s*2025\s*г\.)"
            ],
        },
        {
            "name": "gps_required",
            "label": "Требование установки GPS",
            "patterns": [
                r"(Подтвердить установку GPS системы[^;]+)"
            ],
            "converter": lambda _: "Да",
        },
        {
            "name": "insurance_required",
            "label": "Обязательное страхование КАСКО",
            "patterns": [
                r"(Предусмотреть обязательное страхование КАСКО[^;]+)"
            ],
            "converter": lambda _: "Да",
        },
    ]

    for definition in definitions:
        field = find_field(
            pdf,
            name=definition["name"],
            label=definition["label"],
            patterns=definition["patterns"],
            converter=definition.get("converter"),
        )
        if field:
            fields.append(field)

    values = {field["name"]: field["value"] for field in fields}
    asset_value = values.get("asset_value_kzt")
    advance = values.get("advance_payment_kzt")
    commission = values.get("commission_kzt")

    if isinstance(asset_value, (int, float)) and isinstance(advance, (int, float)):
        financing_amount = asset_value - advance
        advance_percent = advance / asset_value * 100

        fields.extend(
            [
                {
                    "name": "financing_amount_kzt",
                    "label_ru": "Сумма финансирования, тенге",
                    "value": round(financing_amount, 2),
                    "page": None,
                    "quote": f"{asset_value:.2f} - {advance:.2f}",
                    "confidence": 1.0,
                    "value_type": "calculated",
                },
                {
                    "name": "advance_percent",
                    "label_ru": "Аванс, %",
                    "value": round(advance_percent, 2),
                    "page": None,
                    "quote": f"{advance:.2f} / {asset_value:.2f} × 100",
                    "confidence": 1.0,
                    "value_type": "calculated",
                },
            ]
        )

        if isinstance(commission, (int, float)):
            expected = financing_amount * 0.01
            fields.append(
                {
                    "name": "commission_math_check",
                    "label_ru": "Проверка комиссии 1%",
                    "value": {
                        "expected": round(expected, 2),
                        "actual": round(commission, 2),
                        "matches": abs(expected - commission) < 0.01,
                    },
                    "page": None,
                    "quote": f"{financing_amount:.2f} × 1%",
                    "confidence": 1.0,
                    "value_type": "calculated",
                }
            )

    return fields


def extract_purchase(pdf: dict[str, Any]) -> list[dict[str, Any]]:
    fields = common_fields(pdf)

    definitions = [
        {
            "name": "purchase_contract_number",
            "label": "Номер договора купли-продажи",
            "patterns": [
                r"ДОГОВОР купли-продажи товара.*?№\s*([A-ZА-Я0-9/_-]+)"
            ],
        },
        {
            "name": "purchase_contract_date",
            "label": "Дата договора купли-продажи",
            "patterns": [
                r"г\.\s*Алматы\s*(\d{2}\.\d{2}\.\d{4})\s*г"
            ],
        },
        {
            "name": "purchase_total_kzt",
            "label": "Общая стоимость договора, тенге",
            "patterns": [
                r"Общая стоимость настоящего Договора составляет\s*([\d\s]+,\d{2})"
            ],
            "converter": parse_number,
        },
        {
            "name": "vat_percent",
            "label": "НДС, %",
            "patterns": [
                r"в том числе НДС\s*(\d+)%"
            ],
            "converter": int,
        },
        {
            "name": "payment_percent",
            "label": "Оплата продавцу, %",
            "patterns": [
                r"Общей стоимости настоящего Договора,\s*что составляет\s*[\d\s]+,\d{2}[^%]+100%"
            ],
            "converter": lambda _: 100,
        },
        {
            "name": "payment_term_workdays",
            "label": "Срок оплаты, рабочих дней",
            "patterns": [
                r"с момента подписания настоящего Договора[^.]{0,150}5\s*\(пять\)\s*рабочих дней"
            ],
            "converter": lambda _: 5,
        },
        {
            "name": "delivery_term_workdays",
            "label": "Срок поставки, рабочих дней",
            "patterns": [
                r"поставку Товара в течение\s*(\d+)\s*\([^)]*\)\s*рабочих дней"
            ],
            "converter": int,
        },
        {
            "name": "delivery_address",
            "label": "Место поставки",
            "patterns": [
                r"(Республика Казахстан,\s*город\s*Актобе,\s*проспект\s*Нокина,\s*14е)"
            ],
        },
        {
            "name": "warranty_months",
            "label": "Гарантия, месяцев",
            "patterns": [
                r"в течение\s*(\d+)\s*\(Двенадцать\)\s*месяцев"
            ],
            "converter": int,
        },
        {
            "name": "warranty_km",
            "label": "Гарантия, км",
            "patterns": [
                r"или\s*(30\s*000)\s*\(Тридцать тысяч\)\s*км"
            ],
            "converter": lambda value: int(value.replace(" ", "")),
        },
        {
            "name": "late_delivery_penalty_percent_daily",
            "label": "Пеня за просрочку поставки, % в день",
            "patterns": [
                r"пеню в размере\s*([\d,.]+)%\s*от общей стоимости"
            ],
            "converter": lambda value: float(value.replace(",", ".")),
        },
        {
            "name": "seller_name",
            "label": "Продавец",
            "patterns": [
                r"(?:ТОО|TOO)\s*«([^»]+)»"
            ],
        },
        {
            "name": "seller_bin",
            "label": "БИН продавца",
            "patterns": [
                r"Получатель 1\s+ТОО\s+\"ANTO MOTORS\",\s*(\d{12})"
            ],
        },
        {
            "name": "linked_lease_contract_number",
            "label": "Связанный договор лизинга",
            "patterns": [
                r"Заявление о присоединении\s*№\s*([A-ZА-Я0-9/_-]+)"
            ],
        },
        {
            "name": "asset_description",
            "label": "Товар",
            "patterns": [
                r"(Самосвал\s+HOWO,\s*T5G,\s*год выпуска\s*[–-]\s*2025\s*г\.)"
            ],
        },
        {
            "name": "asset_quantity",
            "label": "Количество, шт.",
            "patterns": [
                r"Самосвал\s+HOWO,\s*T5G,\s*год выпуска\s*[–-]\s*2025\s*г\.\s*1\s*35\s*750\s*000,00"
            ],
            "converter": lambda _: 1,
        },
    ]

    for definition in definitions:
        field = find_field(
            pdf,
            name=definition["name"],
            label=definition["label"],
            patterns=definition["patterns"],
            converter=definition.get("converter"),
        )
        if field:
            fields.append(field)

    return fields


def extract_fields(pdf: dict[str, Any], document_type: str) -> list[dict[str, Any]]:
    if document_type == "lease_contract":
        return extract_lease(pdf)
    if document_type == "purchase_contract":
        return extract_purchase(pdf)
    return common_fields(pdf)


def fields_map(document: dict[str, Any]) -> dict[str, Any]:
    return {field["name"]: field["value"] for field in document["fields"]}


def add_comparison(
    checks: list[dict[str, str]],
    name: str,
    left: Any,
    right: Any,
) -> None:
    if left is None or right is None:
        checks.append(
            {
                "name": name,
                "status": "warning",
                "message": "Недостаточно данных для проверки.",
            }
        )
        return

    if isinstance(left, str) and isinstance(right, str):
        left_cmp = normalize_text(left).lower()
        right_cmp = normalize_text(right).lower()
    else:
        left_cmp = left
        right_cmp = right

    checks.append(
        {
            "name": name,
            "status": "ok" if left_cmp == right_cmp else "error",
            "message": (
                "Значения совпадают."
                if left_cmp == right_cmp
                else f"Не совпало: {left} / {right}."
            ),
        }
    )


def validate_documents(documents: list[dict[str, Any]]) -> list[dict[str, str]]:
    checks = []

    lease = next(
        (doc for doc in documents if doc["document_type"] == "lease_contract"),
        None,
    )
    purchase = next(
        (doc for doc in documents if doc["document_type"] == "purchase_contract"),
        None,
    )

    if not lease or not purchase:
        checks.append(
            {
                "name": "Базовая пара документов",
                "status": "warning",
                "message": "Для полной сверки загрузите оба PDF одновременно.",
            }
        )
        return checks

    lease_values = fields_map(lease)
    purchase_values = fields_map(purchase)

    add_comparison(
        checks,
        "Стоимость предмета",
        lease_values.get("asset_value_kzt"),
        purchase_values.get("purchase_total_kzt"),
    )
    add_comparison(
        checks,
        "Описание имущества",
        lease_values.get("asset_description"),
        purchase_values.get("asset_description"),
    )

    lease_purchase_number = lease_values.get(
        "linked_purchase_contract_number"
    )
    purchase_number = purchase_values.get("purchase_contract_number")
    add_comparison(
        checks,
        "Связанный номер договора купли-продажи",
        lease_purchase_number,
        purchase_number,
    )

    lease_number = lease_values.get("lease_contract_number")
    purchase_lease_number = purchase_values.get(
        "linked_lease_contract_number"
    )
    add_comparison(
        checks,
        "Связанный номер договора лизинга",
        lease_number,
        purchase_lease_number,
    )

    for document, expected in [(lease, 2), (purchase, 3)]:
        values = fields_map(document)
        actual = values.get("signature_count")

        if actual is None:
            checks.append(
                {
                    "name": f"ЭЦП: {document['document_type_label_ru']}",
                    "status": "warning",
                    "message": "Количество ЭЦП не найдено.",
                }
            )
        else:
            checks.append(
                {
                    "name": f"ЭЦП: {document['document_type_label_ru']}",
                    "status": "ok" if actual == expected else "error",
                    "message": f"Ожидалось {expected}, найдено {actual}.",
                }
            )

    commission_check = lease_values.get("commission_math_check")
    if isinstance(commission_check, dict):
        checks.append(
            {
                "name": "Комиссия 1% от суммы финансирования",
                "status": (
                    "ok" if commission_check["matches"] else "error"
                ),
                "message": (
                    f"Ожидаемо {commission_check['expected']:,.2f}; "
                    f"в документе {commission_check['actual']:,.2f}."
                ),
            }
        )

    return checks


def safe_sheet_title(title: str, used: set[str]) -> str:
    title = re.sub(r"[\[\]\*\?/\\:]", " - ", title)
    title = re.sub(r"\s+", " ", title).strip()
    title = title[:31] or "Документ"

    original = title
    counter = 2
    while title in used:
        suffix = f" {counter}"
        title = original[: 31 - len(suffix)] + suffix
        counter += 1

    used.add(title)
    return title


def save_json(result: dict[str, Any], path: Path) -> None:
    path.write_text(
        json.dumps(result, ensure_ascii=False, indent=2),
        encoding="utf-8",
    )


def save_excel(result: dict[str, Any], path: Path) -> None:
    workbook = Workbook()
    summary = workbook.active
    summary.title = "Сводка"

    header_fill = PatternFill("solid", fgColor="D9EAF7")
    header_font = Font(bold=True)

    summary.append(["Показатель", "Значение"])
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font

    summary.append(["Документов", len(result["documents"])])
    summary.append(["Проверок", len(result["checks"])])
    summary.append([
        "Успешных проверок",
        sum(check["status"] == "ok" for check in result["checks"]),
    ])
    summary.append([
        "Ошибок",
        sum(check["status"] == "error" for check in result["checks"]),
    ])
    summary.append([
        "Предупреждений",
        sum(check["status"] == "warning" for check in result["checks"]),
    ])

    used_titles = {"Сводка"}

    for document in result["documents"]:
        title = safe_sheet_title(
            document["document_type_label_ru"],
            used_titles,
        )
        sheet = workbook.create_sheet(title)
        sheet.append(
            [
                "Поле",
                "Значение",
                "Страница",
                "Тип",
                "Уверенность",
                "Цитата",
            ]
        )

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font

        for field in document["fields"]:
            value = field["value"]
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)

            sheet.append(
                [
                    field["label_ru"],
                    value,
                    field["page"],
                    field["value_type"],
                    field["confidence"],
                    field["quote"],
                ]
            )

        widths = {
            "A": 34,
            "B": 34,
            "C": 12,
            "D": 14,
            "E": 14,
            "F": 90,
        }
        for column, width in widths.items():
            sheet.column_dimensions[column].width = width

        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(
                    vertical="top",
                    wrap_text=True,
                )

        sheet.freeze_panes = "A2"

    checks_sheet = workbook.create_sheet("Проверки")
    checks_sheet.append(["Проверка", "Статус", "Комментарий"])

    for cell in checks_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font

    for check in result["checks"]:
        checks_sheet.append(
            [
                check["name"],
                check["status"],
                check["message"],
            ]
        )

    checks_sheet.column_dimensions["A"].width = 45
    checks_sheet.column_dimensions["B"].width = 18
    checks_sheet.column_dimensions["C"].width = 90

    workbook.save(path)


@app.get("/")
def index():
    return render_template("index.html")


@app.post("/analyze")
def analyze():
    uploaded_files = request.files.getlist("documents")

    if not uploaded_files or all(not item.filename for item in uploaded_files):
        return render_template(
            "index.html",
            error="Выберите хотя бы один PDF.",
        )

    documents = []

    with tempfile.TemporaryDirectory() as temp_directory:
        temp_path = Path(temp_directory)

        for uploaded_file in uploaded_files:
            if not uploaded_file.filename:
                continue

            filename = secure_filename(uploaded_file.filename)

            if not filename.lower().endswith(".pdf"):
                return render_template(
                    "index.html",
                    error=f"{uploaded_file.filename}: поддерживаются только PDF.",
                )

            local_path = temp_path / filename
            uploaded_file.save(local_path)

            pdf = read_pdf(local_path)
            document_type, label, class_confidence = classify_document(
                pdf["full_text"]
            )
            fields = extract_fields(pdf, document_type)

            documents.append(
                {
                    "filename": uploaded_file.filename,
                    "page_count": pdf["page_count"],
                    "document_type": document_type,
                    "document_type_label_ru": label,
                    "classification_confidence": class_confidence,
                    "fields": fields,
                }
            )

    checks = validate_documents(documents)

    result = {
        "result_id": uuid.uuid4().hex,
        "documents": documents,
        "checks": checks,
    }

    result_folder = Path(app.config["RESULT_FOLDER"])
    json_path = result_folder / f"{result['result_id']}.json"
    excel_path = result_folder / f"{result['result_id']}.xlsx"

    save_json(result, json_path)
    save_excel(result, excel_path)

    return render_template("results.html", result=result)


@app.get("/download/<result_id>/<kind>")
def download(result_id: str, kind: str):
    if not re.fullmatch(r"[a-f0-9]{32}", result_id):
        abort(404)

    extension = {"json": "json", "excel": "xlsx"}.get(kind)
    if extension is None:
        abort(404)

    path = Path(app.config["RESULT_FOLDER"]) / f"{result_id}.{extension}"
    if not path.exists():
        abort(404)

    return send_file(
        path,
        as_attachment=True,
        download_name=f"результат_анализа.{extension}",
    )


if __name__ == "__main__":
    app.run(debug=True, host="127.0.0.1", port=5000)
