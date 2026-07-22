from __future__ import annotations

import json
import re
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def save_json(data: dict, path: Path) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2), encoding='utf-8')


def _safe_sheet_title(title: str, used: set[str]) -> str:
    title = re.sub(r'[\[\]\*\?/\\:]', ' - ', title)
    title = re.sub(r'\s+', ' ', title).strip()[:31] or 'Документ'
    original = title
    counter = 2
    while title in used:
        suffix = f' {counter}'
        title = original[:31-len(suffix)] + suffix
        counter += 1
    used.add(title)
    return title


def save_excel(data: dict, path: Path) -> None:
    wb = Workbook()
    summary = wb.active
    summary.title = 'Сводка'
    header_fill = PatternFill('solid', fgColor='D9EAF7')
    warning_fill = PatternFill('solid', fgColor='FFF2CC')
    header_font = Font(bold=True)

    summary.append(['Показатель', 'Значение'])
    client = data.get("client", {})
    if client:
        summary.append(["Клиент", client.get("name") or "Не определён"])
        summary.append(["ИИН/БИН клиента", client.get("iin_bin") or "Не определён"])
        summary.append(["Роль клиента", client.get("role_label_ru") or "Не определена"])
    for cell in summary[1]:
        cell.fill = header_fill
        cell.font = header_font
    summary.append(['Документов', len(data['documents'])])
    summary.append(['Документов с OCR', sum(bool(doc['used_ocr']) for doc in data['documents'])])
    summary.append(['Неизвестных типов', sum(doc['document_type'] == 'unknown' for doc in data['documents'])])
    summary.append(['Предупреждений', sum(len(doc.get('warnings', [])) for doc in data['documents'])])
    dossier = data.get('dossier') or {}
    if dossier:
        summary.append(['Междокументных совпадений', dossier.get('counts', {}).get('match', 0)])
        summary.append(['Междокументных расхождений', dossier.get('counts', {}).get('mismatch', 0)])
        summary.append(['Не удалось проверить', dossier.get('counts', {}).get('not_enough_data', 0)])
        financial = dossier.get('financial') or {}
        if financial.get('total'):
            summary.append(['Арифметических проверок', financial.get('total', 0)])
            summary.append(['Арифметических расхождений', financial.get('mismatch', 0)])
            summary.append(['Максимальное расхождение, тенге', financial.get('largest_difference_kzt', 0)])

    equipment_tables = [
        (document, table)
        for document in data.get("documents", [])
        for table in document.get("tables", [])
        if table.get("name") == "asset_vin_rows"
    ]
    if equipment_tables:
        summary.append([])
        summary.append(["Техника", "Значение"])
        for cell in summary[summary.max_row]:
            cell.fill = header_fill
            cell.font = header_font
        total_quantity = sum(
            table.get("summary", {}).get("total_quantity") or 0
            for _document, table in equipment_tables
        )
        unique_vins = sum(
            table.get("summary", {}).get("unique_vin_count") or 0
            for _document, table in equipment_tables
        )
        summary.append(["Количество техники, найдено", total_quantity or "Не определено"])
        summary.append(["Уникальных VIN", unique_vins])
        by_type = {}
        for _document, table in equipment_tables:
            for label, quantity in table.get("summary", {}).get("equipment_by_type", {}).items():
                by_type[label] = by_type.get(label, 0) + quantity
        for label, quantity in sorted(by_type.items()):
            summary.append([f"Вид техники: {label}", quantity])

    summary.column_dimensions["A"].width = 42
    summary.column_dimensions["B"].width = 42
    for row in summary.iter_rows():
        for cell in row:
            cell.alignment = Alignment(vertical="top", wrap_text=True)
    summary.freeze_panes = "A2"

    used = {'Сводка'}
    if dossier:
        dossier_sheet = wb.create_sheet('Сверка досье')
        dossier_sheet.append(['Категория', 'Проверка', 'Статус', 'Результат', 'Доказательства'])
        for cell in dossier_sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        status_ru = {'match': 'Совпадает', 'mismatch': 'Расхождение', 'not_enough_data': 'Недостаточно данных'}
        for check in dossier.get('checks', []):
            evidence = '; '.join(
                f"{item.get('filename')} / {item.get('field')} = {item.get('value')}"
                for item in check.get('evidence', [])
            )
            dossier_sheet.append([
                check.get('category'), check.get('check'), status_ru.get(check.get('status'), check.get('status')),
                check.get('message'), evidence,
            ])
        for column, width in {'A':22,'B':42,'C':20,'D':75,'E':100}.items():
            dossier_sheet.column_dimensions[column].width = width
        for row in dossier_sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        dossier_sheet.freeze_panes = 'A2'
        used.add('Сверка досье')
    for document in data['documents']:
        sheet = wb.create_sheet(_safe_sheet_title(document['document_type_label_ru'], used))
        sheet.append(['Поле', 'Категория', 'Значение', 'Проверка', 'Сообщение проверки', 'Исходное значение', 'Страница', 'Метод', 'Уверенность', 'Статус', 'Проверено', 'Примечание', 'Цитата'])
        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
        sorted_fields = sorted(
            document['fields'],
            key=lambda item: (
                item.get('page') is None,
                item.get('page') if item.get('page') is not None else 10**9,
                item.get('label_ru', ''),
            ),
        )
        for item in sorted_fields:
            value = item['value']
            if isinstance(value, (dict, list)):
                value = json.dumps(value, ensure_ascii=False)
            original_value = item.get('original_value')
            if isinstance(original_value, (dict, list)):
                original_value = json.dumps(original_value, ensure_ascii=False)
            validation = item.get("validation") or {}
            sheet.append([
                item['label_ru'], item.get('category'), value,
                "Корректно" if validation.get("valid") else "Требует проверки" if validation else None,
                validation.get("message"),
                original_value, item['page'], item['extraction_method'],
                item['confidence'], item.get('status'), item.get('reviewed_at'),
                item.get('notes'), item['quote'],
            ])
        for column, width in {'A':35,'B':18,'C':42,'D':18,'E':48,'F':42,'G':11,'H':14,'I':13,'J':16,'K':24,'L':38,'M':90}.items():
            sheet.column_dimensions[column].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        sheet.freeze_panes = 'A2'

        for table in document.get('tables', []):
            table_sheet = wb.create_sheet(_safe_sheet_title('Таблица - ' + table.get('label_ru', 'Данные'), used))
            columns = table.get('columns', [])
            table_sheet.append([column.get('label_ru', column.get('key')) for column in columns])
            for cell in table_sheet[1]:
                cell.fill = header_fill
                cell.font = header_font
            for row in table.get('rows', []):
                table_sheet.append([row.get(column.get('key')) for column in columns])
            for column_cells in table_sheet.columns:
                letter = column_cells[0].column_letter
                max_len = max(len(str(cell.value or '')) for cell in column_cells)
                table_sheet.column_dimensions[letter].width = min(max(max_len + 2, 12), 55)
            for row in table_sheet.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(vertical='top', wrap_text=True)
            table_sheet.freeze_panes = 'A2'

        if document.get('warnings'):
            warn_sheet = wb.create_sheet(_safe_sheet_title('Контроль - ' + document['document_type_label_ru'], used))
            warn_sheet.append(['Важность', 'Поле', 'Сообщение'])
            for cell in warn_sheet[1]:
                cell.fill = warning_fill
                cell.font = header_font
            for warning in document['warnings']:
                warn_sheet.append([warning['severity'], warning['field'], warning['message']])
            warn_sheet.column_dimensions['A'].width = 16
            warn_sheet.column_dimensions['B'].width = 35
            warn_sheet.column_dimensions['C'].width = 90

    # Excel sometimes remembers the last generated worksheet as active.
    # Explicitly select the summary so every downloaded workbook opens there.
    wb.active = 0
    for index, worksheet in enumerate(wb.worksheets):
        worksheet.sheet_view.tabSelected = index == 0
    summary.sheet_view.selection[0].activeCell = "A1"
    summary.sheet_view.selection[0].sqref = "A1"
    wb.save(path)
